from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import datetime
from tkinter import *

def tela():

    layout = Tk()
    layout.title("Captador de temperatura SP")
    layout.geometry("200x100")
    layout.resizable(False, False)

    tela = Frame(layout)

    label_texto = Label(tela, text="Aperte para atualizar planilha:")
    atualizar_planilha = Button(tela, text="Atualizar", command=atualizar)

    tela.pack()
    label_texto.pack()
    atualizar_planilha.pack()

    tela.mainloop()

def atualizar():

    navegador = webdriver.Chrome()
    navegador.get('https://br.search.yahoo.com/search?fr=mcafee&type=E210BR826G0&p=previs%C3%A3o+do+tempo+s%C3%A3o+paulo')

    data = datetime.datetime.now()

    temperatura = navegador.find_element(By.XPATH, '//*[@id="left"]/div/ol[1]/li/div/div[1]/div[2]/div/div')
    umidade = navegador.find_element(By.XPATH, '//*[@id="left"]/div/ol[1]/li/div/div[1]/div[2]/ul/li[2]')   
    
    temperatura = temperatura.text
    umidade = umidade.text
        
    arquivo = load_workbook("historico_temperatura.xlsx")
    plan = arquivo['temperatura']
    lista = [data, temperatura, umidade]

    posicao = 0

    for j in range(1, 4):
        linha = True
        i = 2
        while linha:
            if plan.cell(row=i, column=j).value == None:
                plan.cell(row=i, column=j).value = lista[posicao]
                linha = False
            else:
                i = i + 1
        posicao = posicao + 1

    navegador.quit()

    arquivo.save("historico_temperatura.xlsx")

tela()